from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os
import random
import string
from datetime import datetime
from fpdf import FPDF 

app = Flask(__name__)

# Helper function to generate unique Certificate IDs
def generate_certificate_id():
    prefix = "NXSP"
    part1 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    part2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    part3 = prefix + ''.join(random.choices(string.ascii_uppercase + string.digits, k=2))
    return f"{part1}-{part2}-{part3}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_certificates():
    template = request.files['template']
    excel_file = request.files['excel']

    template_path = os.path.join("uploads", template.filename)
    excel_path = os.path.join("uploads", excel_file.filename)
    template.save(template_path)
    excel_file.save(excel_path)

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    students = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        students.append({"Name": row[0]})

    output_folder = "nexospark_certificates"
    os.makedirs(output_folder, exist_ok=True)

    # Use Times New Roman font (adjust the path based on your system)
    font_path = "C:/Windows/Fonts/times.ttf"
    font_name = ImageFont.truetype("C:/Windows/Fonts/timesbd.ttf", 80)
    font_id = ImageFont.truetype(font_path, 15)
    font_time = ImageFont.truetype(font_path, 15)

    # Set grey color for text
    grey_color = (128, 128, 128)  # RGB for grey

    # Loop through each student and generate certificates
    for student in students:
        name = student["Name"]
        cert_id = generate_certificate_id()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Load a fresh copy of the template for each student
        base_image = Image.open(template_path).convert("RGB")
        draw = ImageDraw.Draw(base_image)
        
        name_bbox = draw.textbbox((0, 0), name, font=font_name)
        text_width = name_bbox[2] - name_bbox[0]
        text_height = name_bbox[3] - name_bbox[1]# Width of the text
        name_x = (base_image.width - text_width) // 2  # Center x-coordinate
        name_y = (base_image.height - text_height) // 2 - 200

        # Draw student name
        draw.text((name_x, name_y), f"{name}", fill="black", font=font_name)

        # Create a new image with rotated text for the certificate ID
        id_image = Image.new("RGBA", (300, 50), (255, 255, 255, 0))  # Transparent background
        id_draw = ImageDraw.Draw(id_image)
        id_draw.text((0, 0), cert_id, fill=grey_color, font=font_id)  # Use grey color for ID

        # Rotate the certificate ID 90 degrees counterclockwise
        rotated_id = id_image.rotate(90, expand=True)

        # Paste the rotated ID onto the certificate template
        base_image.paste(rotated_id, (50, -50), rotated_id)  # Adjust (50, 50) as needed

        # Calculate the position for the timestamp (bottom-left corner)
        width, height = base_image.size  # Get the size of the certificate image
        timestamp_bbox = draw.textbbox((0, 0), timestamp, font=font_time)
        timestamp_width = timestamp_bbox[2] - timestamp_bbox[0]
        timestamp_height = timestamp_bbox[3] - timestamp_bbox[1] 
        timestamp_x = 60  # Some padding from the left edge
        timestamp_y = height - 80  # Some padding from the bottom edge

        # Draw the timestamp on the bottom-left corner
        draw.text((timestamp_x, timestamp_y), f"Generated on :{timestamp} +05:30 GMT", fill=grey_color, font=font_time)  # Use grey color for timestamp

        # Save the certificate image temporarily
        temp_image_path = os.path.join(output_folder, f"{name}_certificate.png")
        base_image.save(temp_image_path)

        # Convert the image to PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.image(temp_image_path, x=0, y=0, w=210, h=297)  # A4 size in mm (210x297)
        pdf_path = os.path.join(output_folder, f"{name}_certificate.pdf")
        pdf.output(pdf_path)

        # Write the unique ID in the adjacent cell to the name
        row_index = students.index(student) + 2  # +2 to account for header and zero-indexing
        sheet.cell(row=row_index, column=2).value = cert_id  # Write unique ID in the second column

        # Remove the temporary image to save space
        os.remove(temp_image_path)

    # Save the updated Excel sheet with certificate IDs
    updated_excel_path = os.path.join("uploads", "updated_students.xlsx")
    workbook.save(updated_excel_path)

    return send_file(updated_excel_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
